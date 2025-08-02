from os.path import exists
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

def load_template():
    """Loads Excel template if available"""
    # No Excel template exists in the project, so always return None
    return None

def parse_table(lines, start_idx):
    """Parse markdown table and return the table data and next line index"""
    table_lines = []
    i = start_idx

    # Find all table lines
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith('|') and line.endswith('|'):
            table_lines.append(line)
            i += 1
        else:
            break

    if len(table_lines) < 2:  # Need at least header and separator
        return None, start_idx + 1

    # Parse table data
    table_data = []
    for line in table_lines:
        # Skip separator line (contains dashes)
        if '---' in line or ':-:' in line or ':--' in line or '--:' in line:
            continue

        # Split by | and clean up
        cells = [cell.strip() for cell in line.split('|')[1:-1]]  # Remove empty first/last
        table_data.append(cells)

    return table_data, i

def format_cell_value(value):
    """Convert string value to appropriate Excel type (number, text, formula, etc.)"""
    value = value.strip()

    # Check if it's a formula (starts with =)
    if value.startswith('='):
        return value

    # Try to convert to number
    try:
        # Check for percentage
        if value.endswith('%'):
            return float(value[:-1]) / 100
        # Check for regular number
        return float(value)
    except ValueError:
        return value

def parse_cell_formatting(cell_text):
    """Parse markdown formatting in cell text and return clean text and formatting info"""
    formatting_info = {'bold': False, 'italic': False, 'monospace': False}
    clean_text = cell_text.strip()
    
    # Handle bold text **text**
    if clean_text.startswith('**') and clean_text.endswith('**'):
        clean_text = clean_text[2:-2]
        formatting_info['bold'] = True
    
    # Handle italic text *text* (but not if it was already processed as bold)
    elif clean_text.startswith('*') and clean_text.endswith('*'):
        clean_text = clean_text[1:-1]
        formatting_info['italic'] = True
    
    # Handle code text `text` - use monospace font
    elif clean_text.startswith('`') and clean_text.endswith('`'):
        clean_text = clean_text[1:-1]
        formatting_info['monospace'] = True
    
    return clean_text, formatting_info

def apply_cell_formatting(cell, formatting_info):
    """Apply formatting information to an Excel cell"""
    current_font = cell.font
    
    if formatting_info['bold']:
        cell.font = Font(bold=True, color=current_font.color, size=current_font.size)
    elif formatting_info['italic']:
        cell.font = Font(italic=True, color=current_font.color, size=current_font.size)
    elif formatting_info['monospace']:
        cell.font = Font(name='Courier New', color=current_font.color, size=current_font.size)

def adjust_formula_references(formula, current_excel_row, table_positions=None):
    """Convert row-relative references [offset] and table references T1.B[1] to actual Excel row numbers"""
    if not formula.startswith('='):
        return formula

    if table_positions is None:
        table_positions = {}

    # First handle table-based references like T1.B[1], T2.SUM(C[0]:F[0])
    table_pattern = r'T(\d+)\.([A-Z]+)\[([+-]?\d+)\]'

    def replace_table_reference(match):
        table_num = int(match.group(1))
        column = match.group(2)
        offset = int(match.group(3))

        # Get the starting row of the specified table
        table_key = f"T{table_num}"
        if table_key in table_positions:
            table_start_row = table_positions[table_key]
            # Add 1 to skip header row, then add offset
            actual_row = table_start_row + 1 + offset
            return f"{column}{actual_row}"
        else:
            # Fallback to current table if table not found
            actual_row = current_excel_row + offset
            return f"{column}{actual_row}"

    # Replace table-based cell references
    adjusted_formula = re.sub(table_pattern, replace_table_reference, formula)

    # Handle table-based range references like T1.B[0]:T1.E[0]
    table_range_pattern = r'T(\d+)\.([A-Z]+)\[([+-]?\d+)\]:T(\d+)\.([A-Z]+)\[([+-]?\d+)\]'

    def replace_table_range(match):
        start_table_num = int(match.group(1))
        start_col = match.group(2)
        start_offset = int(match.group(3))
        end_table_num = int(match.group(4))
        end_col = match.group(5)
        end_offset = int(match.group(6))

        # Get starting rows for both tables
        start_table_key = f"T{start_table_num}"
        end_table_key = f"T{end_table_num}"

        if start_table_key in table_positions:
            start_table_row = table_positions[start_table_key]
            start_row = start_table_row + 1 + start_offset
        else:
            start_row = current_excel_row + start_offset

        if end_table_key in table_positions:
            end_table_row = table_positions[end_table_key]
            end_row = end_table_row + 1 + end_offset
        else:
            end_row = current_excel_row + end_offset

        return f"{start_col}{start_row}:{end_col}{end_row}"

    adjusted_formula = re.sub(table_range_pattern, replace_table_range, adjusted_formula)

    # Handle simplified table range references like T1.SUM(B[0]:E[0])
    table_func_pattern = r'T(\d+)\.(SUM|AVERAGE|MAX|MIN)\(([A-Z]+)\[([+-]?\d+)\]:([A-Z]+)\[([+-]?\d+)\]\)'

    def replace_table_function(match):
        table_num = int(match.group(1))
        func_name = match.group(2)
        start_col = match.group(3)
        start_offset = int(match.group(4))
        end_col = match.group(5)
        end_offset = int(match.group(6))

        table_key = f"T{table_num}"
        if table_key in table_positions:
            table_start_row = table_positions[table_key]
            start_row = table_start_row + 1 + start_offset
            end_row = table_start_row + 1 + end_offset
        else:
            start_row = current_excel_row + start_offset
            end_row = current_excel_row + end_offset

        return f"{func_name}({start_col}{start_row}:{end_col}{end_row})"

    adjusted_formula = re.sub(table_func_pattern, replace_table_function, adjusted_formula)

    # Find the current table's start row for relative references
    current_table_start = None
    for table_key, table_start_row in table_positions.items():
        # Check if current_excel_row falls within this table's range
        # We need to find which table contains the current row
        if table_start_row <= current_excel_row:
            current_table_start = table_start_row

    # Finally, handle regular row-relative references [offset] for current table
    pattern = r'([A-Z]+)\[([+-]?\d+)\]'

    def replace_reference(match):
        column = match.group(1)
        offset = int(match.group(2))

        # Calculate from the start of the current table, not the current row
        if current_table_start is not None:
            actual_row = current_table_start + 1 + offset  # +1 to skip header
        else:
            # Fallback to old behavior if we can't determine table start
            actual_row = current_excel_row + offset

        return f"{column}{actual_row}"

    # Replace all remaining row-relative references
    adjusted_formula = re.sub(pattern, replace_reference, adjusted_formula)

    # Handle regular range references like B[0]:E[0] (within current table)
    range_pattern = r'([A-Z]+)\[([+-]?\d+)\]:([A-Z]+)\[([+-]?\d+)\]'

    def replace_range(match):
        start_col = match.group(1)
        start_offset = int(match.group(2))
        end_col = match.group(3)
        end_offset = int(match.group(4))

        # Calculate from the start of the current table, not the current row
        if current_table_start is not None:
            start_row = current_table_start + 1 + start_offset  # +1 to skip header
            end_row = current_table_start + 1 + end_offset
        else:
            # Fallback to old behavior if we can't determine table start
            start_row = current_excel_row + start_offset
            end_row = current_excel_row + end_offset

        return f"{start_col}{start_row}:{end_col}{end_row}"

    adjusted_formula = re.sub(range_pattern, replace_range, adjusted_formula)

    return adjusted_formula

def detect_formula_pattern(value):
    """Detect common formula patterns in markdown and convert to Excel formulas"""
    value = value.strip()

    # Already a formula with relative references or standard Excel formula
    if value.startswith('='):
        return value

    # SUM pattern: SUM(A1:A5) or sum(A1:A5)
    if re.match(r'^(SUM|sum)\([A-Z]+\d+:[A-Z]+\d+\)$', value):
        return f"={value.upper()}"

    # AVERAGE pattern: AVG(A1:A5) or AVERAGE(A1:A5)
    if re.match(r'^(AVG|avg|AVERAGE|average)\([A-Z]+\d+:[A-Z]+\d+\)$', value):
        return f"=AVERAGE({value.split('(')[1]}"

    # Simple arithmetic with cell references: A1+B1, A1*B1, etc.
    if re.match(r'^[A-Z]+\d+[\+\-\*\/][A-Z]+\d+$', value):
        return f"={value}"

    # Percentage calculation: A1/B1*100
    if re.match(r'^[A-Z]+\d+\/[A-Z]+\d+\*100$', value):
        return f"={value}/100"  # Convert to decimal for percentage format

    return value

def add_table_to_sheet(table_data, worksheet, start_row, table_positions=None):
    """Add table data to Excel worksheet with proper formatting and formula support"""
    if not table_data:
        return start_row

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    formula_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Add table data
    for row_idx, row_data in enumerate(table_data):
        current_excel_row = start_row + row_idx

        for col_idx, cell_text in enumerate(row_data):
            cell = worksheet.cell(row=current_excel_row, column=col_idx + 1)

            # First, parse markdown formatting to get clean text and formatting info
            clean_text, formatting_info = parse_cell_formatting(cell_text)

            # Detect and format formulas using the clean text
            formula_value = detect_formula_pattern(clean_text)

            # Format cell value (convert numbers, percentages, formulas)
            if formula_value.startswith('='):
                # Adjust row-relative references to actual Excel rows
                adjusted_formula = adjust_formula_references(formula_value, current_excel_row, table_positions)
                cell.value = adjusted_formula
                cell.fill = formula_fill  # Light blue background for formulas
            else:
                formatted_value = format_cell_value(clean_text)
                cell.value = formatted_value

            # Apply markdown formatting (bold, italic, code) if any was detected
            apply_cell_formatting(cell, formatting_info)

            cell.border = border

            # Set alignment
            if row_idx == 0:
                cell.alignment = Alignment(horizontal='center')
            elif isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith('=')):
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')

            # Style header row
            if row_idx == 0:
                cell.font = header_font
                cell.fill = header_fill
            elif isinstance(cell.value, float) and 0 <= cell.value <= 1 and not cell.value == 0:
                # Format as percentage if it's a decimal between 0 and 1
                cell.number_format = '0.00%'
            elif isinstance(cell.value, (int, float)) and cell.value >= 1000:
                # Format large numbers with thousands separator
                cell.number_format = '#,##0'

    # Auto-adjust column widths
    for col_idx in range(len(table_data[0]) if table_data else 0):
        column_letter = get_column_letter(col_idx + 1)
        max_length = 0
        for row in table_data:
            if col_idx < len(row):
                max_length = max(max_length, len(str(row[col_idx])))
        adjusted_width = min(max(max_length + 2, 12), 25)  # Min 12, max 25 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width

    return start_row + len(table_data) + 2  # Return next available row with spacing

def markdown_to_excel(markdown_content):
    """Convert Markdown to Excel workbook (focused on tables and headers)."""
    template_path = load_template()

    # Create workbook
    if template_path:
        try:
            wb = load_workbook(template_path)
            ws = wb.active
        except Exception as e:
            print(f"Warning: Could not load template {template_path}: {e}")
            wb = Workbook()
            ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    # Set worksheet title
    ws.title = "Data Report"

    # Split content into lines
    lines = markdown_content.split('\n')
    current_row = 1
    table_counter = 1
    table_positions = {}  # Track where each table starts
    i = 0

    try:
        while i < len(lines):
            line = lines[i].strip()

            # Skip empty lines
            if not line:
                i += 1
                continue

            # Headers
            if line.startswith('#'):
                header_level = len(line) - len(line.lstrip('#'))
                header_text = line.lstrip('#').strip()

                cell = ws.cell(row=current_row, column=1)
                cell.value = header_text

                # Style headers based on level
                if header_level == 1:
                    cell.font = Font(size=16, bold=True, color="2F5597")
                elif header_level == 2:
                    cell.font = Font(size=14, bold=True, color="4472C4")
                else:
                    cell.font = Font(size=12, bold=True)

                current_row += 2  # Add space after headers
                i += 1

            # Tables
            elif line.startswith('|'):
                table_data, i = parse_table(lines, i)
                if table_data:
                    # Record this table's position
                    table_key = f"T{table_counter}"
                    table_positions[table_key] = current_row

                    # Process the table
                    current_row = add_table_to_sheet(table_data, ws, current_row, table_positions)
                    table_counter += 1

            # Skip other content
            else:
                i += 1

    except Exception as e:
        print(f"Error in parsing markdown: {e}")
        import traceback
        traceback.print_exc()
        return f"Error in parsing markdown: {e}"

    # Save the workbook to BytesIO and upload
    try:
        from upload_file import upload_file
        import io

        # Save to BytesIO object
        file_object = io.BytesIO()
        wb.save(file_object)
        file_object.seek(0)

        # Upload and get result
        result = upload_file(file_object, "xlsx")
        file_object.close()

        print(f"Excel document uploaded successfully")
        return result
    except Exception as e:
        print(f"Error saving/uploading Excel document: {e}")
        import traceback
        traceback.print_exc()
        return f"Error saving/uploading Excel document: {e}"
